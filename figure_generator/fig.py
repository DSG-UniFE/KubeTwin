import pandas as pd
import numpy as np
import seaborn as sns; sns.set()
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import sys

if len(sys.argv) < 2:
    print("fig.py <csv_file>")
    sys.exit(1)

df = pd.read_csv(sys.argv[1])
grouped = df.groupby('Component')['TTP']
groupedTime = df.groupby('Component')['Time']



wb = load_workbook(filename='service_time.xlsx', 
                   read_only=True)
ws = wb.active
rangeServices = ws['A2':'A100']
rangeIotDG = ws['B2':'B100']

servicesValues = []
iotDGValues = []

for cell in rangeServices:
    for x in cell:
        servicesValues.append(x.value)

for cell in rangeIotDG:
    for x in cell:
        iotDGValues.append(x.value)



fig, ax = plt.subplots(3, 1, figsize=(18,10), sharey=True, squeeze= False)
plt.subplots_adjust(left=0.125, bottom=0.1, right=0.9, top=0.9, wspace=0.2, hspace=0.4)

i = 0 #index used in subplots generation

for key, grp in df.groupby(['Component']):
    # calculate SMA for TTP
    grp['TTP_SMA'] = grp['TTP'].rolling(window=5).mean()
    ax[i,0].plot(grp['Time'], grp['TTP_SMA'], label='Simulated TTP')
    i += 1



fig.text(0.5, 0.04, 'Time (s)', ha='center', fontsize='large')
fig.text(0.04, 0.5, 'TTP (s)', va='center', rotation='vertical', fontsize='large')

ax[0,0].set_title('IoT Client')
ax[1,0].set_title('IoT data Generator')
ax[2,0].set_title('Service List')

ax[1,0].plot(groupedTime.get_group('IoT data Generator'), iotDGValues, color='green', label='Real TTP')
ax[2,0].plot(groupedTime.get_group('Service List'), servicesValues, color='green', label='Real TTP')




lines, labels = fig.axes[-1].get_legend_handles_labels()
fig.legend(lines, labels, loc = 'upper left', fontsize='large')


plt.show()
fig.savefig('fig1.pdf')

fig, ax = plt.subplots(figsize=(10,4))
for key, grp in df.groupby(['Component']):
    ax.plot(grp['Time'],grp['Pods'],label=key)

ax.legend()
plt.xlabel('time (s)')
plt.ylabel('# Pods')
#plt.show()
#fig.savefig('fig2.pdf')



